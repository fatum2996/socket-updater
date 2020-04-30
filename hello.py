#функция, ищущая фото которых нет на сайте
#функция, ищущая EXTRA фото которых нет на сайте
#функцию проверки .JPG

def find_partname(header, header_name):
	for item in header:
		if item.value == header_name:
			return item.column	
	return 0

def load_partnames_to_set(ws, header_name):
	result = set()
	row_header = ws[1]
	c = find_partname(row_header, header_name)
	if(c != 0):
		for cell in ws[xlsxwriter.utility.xl_col_to_name(c-1)]:			
			result.add(cell.value)		
	else:
		print("No Partname column")
	return result

def array_to_file(filename, array):
	with open(filename, "w") as file:
		for line in array:
			file.write(line+"\n")

def add_missing_to_array(set1,set2):
	list1 = list()
	for item in set1:
		if not (item in set2):
			list1.append(item)
	list1.sort()
	return list1

def find_package(ws, partname, type1): 
	if partname[4:6] == "AD":
		c = 217
	else:
		c = 2
	for row in ws.values:
		if row[1] == type1 and row[2] == c:
				return row[0]
	return type1

def find_manufacturer(ws, manufacturer):
	for row in ws.values:
		if(row[1] == manufacturer):
			return row[0]
	return manufacturer

def ready_to_publish(socket):		
	if not socket.partname:
		return "Socket partname is missing"
	if not socket.pin_count:
		return "Socket pin count is missing"
	if not socket.package_type:
		return "Socket package_type is missing"
	if not socket.pitch:
		return "Socket pitch is missing"
	if not socket.length:
		return "Socket length is missing"
	if not socket.width:
		return "Socket width is missing"
	if not socket.socket_name:
		return "Socket socket_name is missing"
	if not socket.manufacturer:
		return "Socket manufacturer is missing"
	if not socket.drawing_exists:
		return "Socket drawing is missing"			
	if not socket.supplier:
		return "Socket socket.supplier is missing"		
	if socket.manufacturer == "Тест-Контакт":
		return "Socket manufacturer is test-kontakt"						
	return "OK"	

def pick_photos(socket):
	from shutil import copy
	if(socket.main_photo != "Photos/no-image.png"):
		copy("./"+socket.main_photo, "./PickedPhotos/")
	if socket.extra_photos != None:
		extra_list = socket.extra_photos.split(", ")
		for item in extra_list:
			copy("./"+item, "./PickedPhotos/")

class Socket:
	id1 = 0
	partname = ""
	main_photo = ""
	photo_exists = False
	amount = 0
	location = 0 
	package_type = ""
	pin_count = 0
	pitch = 0
	length = 0
	width = 0
	height = 0
	package_drawing = ""
	package_drawing_exists = False
	socket_name = ""
	manufacturer = ""
	drawing = ""
	drawing_exists = False
	dwg = ""
	supplier = ""
	description = ""
	extra_photos = ""
	to_publish = False
	def __init__(self, row):
		self.id1 = row[0]
		self.partname = row[1]
		self.main_photo =  row[2]
		if(row[2] == "Photos/no-image.png"):
			self.photo_exists = False
		if(row[3] != None):
			self.amount = int(row[3])
			self.location = row[4]
		self.package_type = row[5]
		self.pin_count = row[6]
		self.pitch = row[7]
		self.length = row[8]
		self.width = row[9]
		self.height = row[10]
		if(row[11] != ""):
			self.package_drawing = row[11]
			self.package_drawing_exists = True
		self.socket_name = row[12]
		self.manufacturer = row[13]
		if(row[14] != ""):
			self.drawing = row[14]
			self.drawing_exists = True
		self.dwg = row[15]
		self.supplier = row[17]
		self.description = row[18]
		self.extra_photos = row[20]		
		if row[21]:
			if row[21] != "False":
				self.to_publish = True
			else:
				self.to_publish = False
		else:
			self.to_publish = False
	def publish(self, ws, ws_packages, ws_manufacturers):
		from copy import copy
		from openpyxl.styles import PatternFill
		max = ws.max_row
		ws.cell(row = max + 1, column = 1, value = "??") #ID		
		ws.cell(row = max + 1, column = 2, value = self.partname) #Наименование
		ws.cell(row = max + 1, column = 2).fill = PatternFill("solid", fgColor="FFFF00")
		ws.cell(row = max + 1, column = 3, value = self.main_photo) #Картинка
		ws.cell(row = max + 1, column = 4, value = self.extra_photos) #ExtraPhotos
		ws.cell(row = max + 1, column = 5, value = self.partname) #Полное наименование
		ws.cell(row = max + 1, column = 6, value = self.partname.lower()) #url
		ws.cell(row = max + 1, column = 7, value = self.description) #Описание
		ws.cell(row = max + 1, column = 8, value = self.partname+" "+str(self.package_type)+"-"+str(self.pin_count)+" "+str(self.package_type)+" "+str(self.pin_count)) #Аннотация
		ws.cell(row = max + 1, column = 10, value = 2) #Шаблон
		ws.cell(row = max + 1, column = 11, value = find_package(ws_packages, self.partname, self.package_type)) #заменить на функцию #Тип корпуса
		ws.cell(row = max + 1, column = 12, value = self.amount) #Amount
		ws.cell(row = max + 1, column = 13, value = self.pin_count) #Pin Count
		ws.cell(row = max + 1, column = 14, value = self.pitch) #Pitch
		ws.cell(row = max + 1, column = 15, value = self.length) #Length
		ws.cell(row = max + 1, column = 16, value = self.width) #Width
		ws.cell(row = max + 1, column = 17, value = self.height) #Height
		ws.cell(row = max + 1, column = 18, value = self.socket_name) #Socket name
		ws.cell(row = max + 1, column = 19, value = find_manufacturer(ws_manufacturers, self.manufacturer))#заменить на функцию #Socket manufacturer
		ws.cell(row = max + 1, column = 24, value = 1) #Published		
		for i in range(1, 25):
			ws.cell(row = max + 1, column = i).font = copy(ws.cell(row = 3, column = i).font)
			ws.cell(row = max + 1, column = i).alignment = copy(ws.cell(row = 3, column = i).alignment) 
def yes_or_no(question):
    reply = str(input(question+' (y/n): ')).lower().strip()
    if reply[0] == 'y':
        return True
    if reply[0] == 'n':
        return False
    else:
        return yes_or_no("Uhhhh... please enter ")

from openpyxl import load_workbook
import xlsxwriter 
from shutil import copyfile
wb_base = load_workbook(filename = "bd.xlsx")
ws_base = wb_base["Контактирующие"]
wb_online = load_workbook(filename = "sockets.xlsx")
ws_online = wb_online["sockets"] #загружаем базу и онлайн
wb_packages = load_workbook(filename = "category.xlsx")
ws_packages = wb_packages["category_20190402"] #загружаем категории
wb_manufacturers = load_workbook(filename = "manufacture.xlsx") #загружаем производителей
ws_manufacturers = wb_manufacturers["manufacture"]
described = load_partnames_to_set(ws_base, "Partname") #создаем множество описанных КУ
loaded = load_partnames_to_set(ws_online, "Наименование") #создаем множество загруженных КУ
not_online = add_missing_to_array(described, loaded) #делаем список не загруженных
not_online.remove("Partname") #удаляем заголовок
array_to_file("not_online.txt", not_online) #пишем в файл
array_to_file("wrong_online.txt", add_missing_to_array(loaded, described)) #пишем лишние из загруженных в файл


to_import = []
for item in not_online: #все незагруженные
	for row in ws_base.values: #ищем в БД незагруженные
		if(row[1] == item): #если нашли
			socket_to_add = Socket(row) #создаем из ряда объект
			if ready_to_publish(socket_to_add) == "OK" and socket_to_add.to_publish:
				to_import.append(socket_to_add)
				print(socket_to_add.partname)
				break
			#else:
				#if ready_to_publish(socket_to_add) != "OK":
				#	print(socket_to_add.partname + " is not ready to publish: " + ready_to_publish(socket_to_add))
				#else:
				#	print(socket_to_add.partname + " is not banned to publish in bd")
if yes_or_no("Publish?"):
	for socket in to_import:
		pick_photos(socket)
		socket.publish(ws_online, ws_packages, ws_manufacturers)
		print(socket.partname + " is published")			
max = ws_online.max_row
wb_online.save("sockets.xlsx")